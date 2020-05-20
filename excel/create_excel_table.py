'''
Created on Apr 30, 2020

@author: Michael Amos
'''
# Github Page: 
# https://github.com/mamos88/python

# Documentation: 
# https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import openpyxl

FILE_NAME = r"C:\Users\go_ma\eclipse-workspace\python\excel\Football Scores.xlsx"

def create_table(sheet, table_name):
    """Creates a table, sets filters and add column and row stripes"""
    max_rows = sheet.max_row
    max_columns = sheet.max_column
    tab = Table(displayName=table_name,
                ref=f"A1:{get_column_letter(max_columns)}{max_rows}")
    
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    return sheet


""" Load Workbook """
WB = openpyxl.load_workbook(FILE_NAME)

""" Apply formatting to Football Scores sheet"""
SHEET = WB['Football Scores']
SHEET = create_table(SHEET, "Football_Table")


WB.save(FILE_NAME)

print("Finished")
