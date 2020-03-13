import openpyxl
from collections import defaultdict

# Documentation: https://openpyxl.readthedocs.io/en/stable/

file_name = r"C:\Users\go_ma\eclipse-workspace\python\excel\foosball.xlsx"


def populate_dicts(my_sheet, my_dict):
    """Function to read data and populate results dictionaries"""
    sorted_key_list = (sorted(my_dict, key=my_dict.get, reverse=True))
    rownum = 2
    for key in sorted_key_list:
        my_sheet["A" + str(rownum)].value = key
        my_sheet["B" + str(rownum)].value = my_dict[key]
        rownum += 1


def create_sheet(sheet_name, index, header_list):
    """ Function to create sheet and set first row with header values """
    my_sheet = WB.create_sheet(sheet_name, index)
    for idx, header in enumerate(header_list):
        my_sheet.cell(row=1, column=idx + 1).value = header
    return my_sheet


""" Load Workbook """
WB = openpyxl.load_workbook(file_name)

""" Open Sheet """
SHEET = WB['Foosball Data']

""" Create Objects to Store Results """
PLAYER_DICT = defaultdict(int)
SHOT_DICT = defaultdict(int)

""" Read data from columns, count players and goals scored """
for row in range(2, SHEET.max_row + 1):
    player = SHEET["A" + str(row)].value
    shot = SHEET["B" + str(row)].value
    if len(shot) <= 1:
        shot = "Unknown"
    goal = SHEET["C" + str(row)].value
    if goal:
        PLAYER_DICT[player] += 1
        SHOT_DICT[shot] += 1

""" Create Worksheet """
PLAYER_SHEET = create_sheet("Player Sheet", 1, ["player", "Goals Scored"])
SHOT_SHEET = create_sheet("Shot Sheet", 2, ["Shot", "Goals Scored"])

""" Populate dictionaries """
populate_dicts(PLAYER_SHEET, PLAYER_DICT)
populate_dicts(SHOT_SHEET, SHOT_DICT)

WB.save(file_name)

print("Finished")
